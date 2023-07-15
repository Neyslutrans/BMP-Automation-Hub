from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles import numbers
import openpyxl

def format_sheet(sheet):
    try:
        sheet.auto_filter.ref = sheet.dimensions
        sheet.freeze_panes = 'A2'

        for cell in sheet[1]:
            cell.fill = PatternFill(start_color="5F4876", end_color="5F4876", fill_type="solid")
            cell.font = Font(color="FFFFFF")
            cell.alignment = Alignment(horizontal='left')

        percent_format = numbers.FORMAT_PERCENTAGE_00
        for cell in sheet['E']:
            cell.number_format = percent_format

    except Exception as e:
        raise Exception("Ошибка при форматировании рабочего листа") from e

def create_workbook():
    return openpyxl.Workbook()

def save_workbook(workbook, name):
    del workbook['Sheet']
    workbook.save(name)
