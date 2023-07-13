from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles import numbers
import pandas as pd
import openpyxl
import re

from modules import paths
from typing import Dict, Any


def main() -> None:
    try:
        # Выбираем файлы с помощью диалогового окна
        file_paths = paths.PathManager.open_file_dialog()

        # Создаем рабочую книгу
        wb = openpyxl.Workbook()

        counter = 1

        # Обрабатываем файлы
        for file in file_paths:
            try:
                print(f'Обрабатывается файл: {file}')
                
                # Импортируем файл
                df = pd.read_csv(file, encoding='cp1251', sep='\t', header=None, low_memory=False)

                df = process_dataframe(df)

                # Создаем рабочий лист
                ws = wb.create_sheet(f'{df["target"]} - {counter}')
                counter += 1

                # Импортируем датафрейм на рабочий лист
                for row in dataframe_to_rows(df["dataframe"], index=False, header=True):
                    ws.append(row)

                # Форматируем лист Excel
                format_sheet(ws)

            except Exception as e:
                print(f"Ошибка при обработке файла: {file}")
                print(f"Тип ошибки: {type(e).__name__}")
                print(f"Сообщение об ошибке: {str(e)}")
                continue

        # Определяем имя файла для сохранения с помощью диалогового окна
        name = paths.PathManager.save_file_dialog()

        # Сохраняем Excel книгу
        del wb['Sheet']
        wb.save(name)

    except Exception as e:
        print("Произошла ошибка во время выполнения программы.")
        print(f"Тип ошибки: {type(e).__name__}")
        print(f"Сообщение об ошибке: {str(e)}")


def process_dataframe(df: pd.DataFrame) -> Dict[str, Any]:
    try:
        df.drop(df.columns[[0, 4]], axis=1, inplace=True)

        # Определяем целевую аудиторию
        target = df.iloc[0, 3]

        # Переименовываем столбцы
        df.columns = ['Блок распространения', 'Телеканал', 'Позиция', target]
        df.drop([0, 1], axis=0, inplace=True)

        # Форматируем телеканалы и определяем формат чисел
        df['Телеканал'] = df['Телеканал'].apply(lambda x: get_cleared_channel(x))
        df[target] = df[target].apply(lambda x: str(x).replace(',', '.').replace('\xa0', '')).astype('float64')

        # Округляем значения до 4 знаков
        df = df.round(4)

        return {
            'dataframe': df,
            'target': target
        }

    except Exception as e:
        raise Exception("Ошибка при обработке датафрейма") from e


def get_cleared_channel(string: str) -> str:
    '''Возвращает строку с телеканалом без скобок и значений в них'''
    return re.sub(r'\([^()]*\)', '', string).strip()


def format_sheet(sheet: openpyxl.worksheet.worksheet) -> None:
    '''Форматирует рабочий лист Excel книги'''
    try:
        sheet.auto_filter.ref = sheet.dimensions
        sheet.freeze_panes = 'A2'

        for cell in sheet[1]:
            cell.fill = PatternFill(start_color="5F4876", end_color="5F4876", fill_type="solid")
            cell.font = Font(color="FFFFFF")
            cell.alignment = Alignment(horizontal='left')

        percent_format = numbers.FORMAT_PERCENTAGE_00
        for cell in sheet['E']:
            cell.number_format = percent_format

    except Exception as e:
        raise Exception("Ошибка при форматировании рабочего листа") from e


if __name__ == '__main__':
    main()
