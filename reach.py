from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles import numbers
import pandas as pd
import openpyxl
import re

from modules import paths
from typing import Dict, Any


def main() -> None:
    try:
        # Выбираем файлы с помощью диалогового окна
        file_paths = paths.PathManager.open_file_dialog()

        # Создаем рабочую книгу
        wb = openpyxl.Workbook()

        counter = 1

        # Обрабатываем файлы
        for file in file_paths:
            try:
                print(f'Обрабатывается файл: {file}')
                
                # Импортируем файл
                df = pd.read_csv(file, encoding='cp1251', sep='\t', header=None, low_memory=False)

                df = process_dataframe(df)

                # Создаем рабочий лист
                ws = wb.create_sheet(f'{df["target"]} - {counter}')
                counter += 1

                # Импортируем датафрейм на рабочий лист
                for row in dataframe_to_rows(df["dataframe"], index=False, header=True):
                    ws.append(row)

                # Форматируем лист Excel
                format_sheet(ws)

            except Exception as e:
                print(f"Ошибка при обработке файла: {file}")
                print(f"Тип ошибки: {type(e).__name__}")
                print(f"Сообщение об ошибке: {str(e)}")
                continue

        # Определяем имя файла для сохранения с помощью диалогового окна
        name = paths.PathManager.save_file_dialog()

        # Сохраняем Excel книгу
        del wb['Sheet']
        wb.save(name)

    except Exception as e:
        print("Произошла ошибка во время выполнения программы.")
        print(f"Тип ошибки: {type(e).__name__}")
        print(f"Сообщение об ошибке: {str(e)}")


def process_dataframe(df: pd.DataFrame) -> Dict[str, Any]:
    
    REMAINS = 50

    try:
        df.drop(df.columns[[0, 3]], axis=1, inplace=True)

        target = df.iloc[0, 2]
        header = ['Рекламодатель', 'Дата', f'{target} TVR']

        for i in range(3, len(df.columns)):
            header.append(df.iloc[2, i])

        df.columns = header

        df.drop([0, 1, 2], axis=0, inplace=True)
        df.reset_index(drop=True, inplace=True)

        for column in header[2:]:
            df[column] = df[column].apply(lambda x: x.replace(',', '.')).astype('float64')

        df['Дата'] = df['Дата'].apply(lambda x: x.replace('/', '.'))

        df['Total TVR'] = None

        df.loc[0, 'Total TVR'] = df.loc[0, f'{target} TVR']

        for i in range(1, len(df['Total TVR'])):
                df.loc[i, 'Total TVR'] = df.loc[i, f'{target} TVR'] + df.loc[i - 1, 'Total TVR']
                

        df['Остаток'] = df['Total TVR'] % REMAINS
        df['Шаг'] = df['Total TVR'] - df['Остаток']

        df['Отсечка'] = None

        for i in range(1, len(df['Остаток'])):
                df.loc[i, 'Отсечка'] = 'yes' if df.loc[i - 1, 'Остаток'] > df.loc[i, 'Остаток'] else 'no'

        df.loc[len(df) - 1, 'Отсечка'] = 'yes'

        # Округляем значения до 4 знаков
        df = df.round(4)

        return {
            'dataframe': df,
            'target': target
        }

    except Exception as e:
        raise Exception("Ошибка при обработке датафрейма") from e


def format_sheet(sheet: openpyxl.worksheet.worksheet) -> None:
    '''Форматирует рабочий лист Excel книги'''
    try:
        sheet.auto_filter.ref = sheet.dimensions
        sheet.freeze_panes = 'A2'

        for cell in sheet[1]:
            cell.fill = PatternFill(start_color="5F4876", end_color="5F4876", fill_type="solid")
            cell.font = Font(color="FFFFFF")
            cell.alignment = Alignment(horizontal='left')

        percent_format = numbers.FORMAT_PERCENTAGE_00
        for cell in sheet['E']:
            cell.number_format = percent_format

    except Exception as e:
        raise Exception("Ошибка при форматировании рабочего листа") from e


if __name__ == '__main__':
    main()
